import sys
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font
import os

def generar_reporte(ruta_csv):
    df = pd.read_csv(ruta_csv)
    df["archivo_origen"] = os.path.basename(ruta_csv)

    os.makedirs("output", exist_ok=True)
    nombre = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta_out = os.path.join("output", nombre)

    with pd.ExcelWriter(ruta_out, engine="openpyxl") as writer:
        # Hoja 1: datos completos
        df.to_excel(writer, sheet_name="Datos", index=False)

        # Hoja 2: resumen
        resumen = df.groupby("archivo_origen").agg(
            total_filas=("archivo_origen", "count")
        ).reset_index()
        resumen.to_excel(writer, sheet_name="Resumen", index=False)

        # Formato encabezado
        ws = writer.sheets["Datos"]
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    print(f"REPORTE_OK:{ruta_out}")  # Power Automate lee este output

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("ERROR: falta ruta del CSV")
        sys.exit(1)
    generar_reporte(sys.argv[1])
